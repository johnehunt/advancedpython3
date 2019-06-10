from openpyxl import Workbook
import fintech.accounts as accounts


def write_account_transaction_to_excel(filename, account):
    print('Starting write of Excel example')
    workbook = Workbook()
    # Get the current active worksheet
    ws = workbook.active
    ws.title = 'transactions'

    ws['A1'] = 'transaction type'
    ws['B1'] = 'amount'

    row = 2

    # Write out the transactions
    for transaction in account.history:
        ws['A' + str(row)] = transaction.action
        ws['B' + str(row)] = transaction.amount
        row += 1

    workbook.save(filename)

    print('Done Write Excel Example')


print('Starting')
acc = accounts.CurrentAccount('123', 'John', 10.05, 100.0)
acc.deposit(23.45)
acc.withdraw(12.33)

print('Writing Account Transactions')
write_account_transaction_to_excel('accounts.xlsx', acc)

print('Done')

from openpyxl import Workbook


print('Starting Write Excel Example with openPyXL')

# Create an empty workbook
workbook = Workbook()
# Get the current active worksheet
sheet = workbook.active
sheet.title = 'dataset1'
sheet.sheet_properties.tabColor = '0F45F7'

sheet['A1'] = 'John'
sheet['B1'] = 42
sheet['C1'] = 56
sheet['D1'] = '=SUM(B1,C1)/2'
sheet['A2'] = 'Adam'
sheet['B2'] = 75
sheet['C2'] = 86
sheet['D2'] = '=SUM(B2, C2)/2'

sheet2 = workbook.create_sheet(title='dataset2')
sheet2.append(['John', 1, 2, 3, '=SUM(B1,D1)/3'])
sheet2.append(['Adam', 5, 7, 8])

sheet2.cell(column=1, row=3, value='John')
sheet2.cell(column=2, row=3, value=15)


workbook.save('sample2.xlsx')

print('Done Write Excel Example')

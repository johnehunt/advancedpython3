""" Example illustrating reading EXCEL files using the
 openpyxl library. """

from openpyxl import load_workbook


def main():
    print('Starting reading Excel file using openPyXL')

    # data_only = True loads cached reuslt of equations
    # data_only = False load equations (default)
    workbook = load_workbook(filename='sample.xlsx', data_only=True)
    print(workbook.active)
    print(workbook.sheetnames)
    print(workbook.worksheets)

    # print('-' * 10)
    ws = workbook['my worksheet']
    # print(ws['A1'].value)
    # print(ws['B1'].value)

    # print('-' * 10)
    # for sheet in workbook:
    #     print(sheet.title)

    print('-' * 10)
    # print(ws.max_column)
    # print(ws.max_row)

    cell_range = ws[ws.calculate_dimension()]
    for cell in cell_range:
        try:
            print(cell[0].value, end=", ")
            age = int(cell[1].value)
            if age > 18:
                print('You can drink')
            else:
                print('Underage')
        except ValueError:
            print('Error in age data')
            print(cell)
        print(cell[2].value, cell[3].value, cell[4].value)

    print('-' * 10)

    print('Finished reading Excel file using openPyXL')


if __name__ == '__main__':
    main()
